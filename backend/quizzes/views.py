import csv
import io
import json
from pathlib import Path

from django.db.models import Count, Q
from rest_framework import permissions, serializers, status, viewsets
from rest_framework.decorators import action
from rest_framework.response import Response

from subjects.models import Module, Subject, Topic

from .models import Question, QuestionBank, Quiz, QuizQuestion
from .serializers import QuestionBankSerializer, QuizSerializer


class IsTeacherOrOwner(permissions.BasePermission):
    def has_permission(self, request, view):
        user = request.user
        return bool(user and user.is_authenticated and (user.is_teacher() or user.is_admin()))

    def has_object_permission(self, request, view, obj):
        owner_id = getattr(obj, 'owner_id', None)
        if owner_id is None and hasattr(obj, 'created_by_id'):
            owner_id = obj.created_by_id
        return owner_id == request.user.id or request.user.is_admin()


class QuizViewSet(viewsets.ModelViewSet):
    queryset = Quiz.objects.all().order_by('-created_at')
    serializer_class = QuizSerializer

    def get_queryset(self):
        qs = (
            Quiz.objects
            .select_related('owner', 'subject_ref', 'topic_ref', 'module_ref')
            .prefetch_related(
                'questions__choices',
                'assigned_fields',
                'quiz_question_links__question__choices',
                'quiz_question_links__question_bank_reference',
            )
            .annotate(question_count=Count('questions', distinct=True))
            .order_by('-created_at')
        )
        request = self.request
        user = request.user

        q = (request.query_params.get('search') or request.query_params.get('q') or '').strip()
        category = (request.query_params.get('category') or '').strip()
        folder = (request.query_params.get('folder') or '').strip()
        visibility = (request.query_params.get('visibility') or '').strip().lower()
        sort = (request.query_params.get('sort') or '').strip().lower()
        subject = (request.query_params.get('subject') or '').strip()
        subject_id = (request.query_params.get('subject_id') or '').strip()
        topic_id = (request.query_params.get('topic_id') or '').strip()
        module_id = (request.query_params.get('module_id') or '').strip()
        quiz_type = (request.query_params.get('quiz_type') or '').strip().lower()
        difficulty = (request.query_params.get('difficulty') or '').strip().lower()
        field_id = (request.query_params.get('field') or request.query_params.get('field_id') or '').strip()
        semester_code = (request.query_params.get('semester_code') or '').strip()
        semester = (request.query_params.get('semester') or '').strip()
        semester_number = (request.query_params.get('semester_number') or '').strip()
        section = (request.query_params.get('section') or '').strip()
        is_published = (request.query_params.get('is_published') or '').strip().lower()
        min_questions = request.query_params.get('min_questions')
        mine = (request.query_params.get('mine') or '').strip().lower()

        if user and user.is_authenticated:
            if getattr(user, 'is_staff', False) or (hasattr(user, 'is_admin') and user.is_admin()):
                pass
            elif hasattr(user, 'is_teacher') and user.is_teacher():
                qs = qs.filter(Q(owner=user) | Q(is_published=True))
            elif hasattr(user, 'is_student') and user.is_student():
                student_field_id = getattr(user, 'field_of_study_id', None)
                student_semester_code = str(getattr(user, 'semester_code', '') or '').strip()
                student_semester_number = getattr(user, 'semester_number', None)
                student_section = str(getattr(user, 'section', '') or '').strip()
                qs = qs.filter(is_published=True)
                visibility_filter = Q(apply_to_all_fields=True)
                if student_field_id:
                    visibility_filter |= Q(assigned_fields__id=student_field_id)
                    visibility_filter |= Q(target_field_of_study_id=student_field_id)
                else:
                    visibility_filter &= Q(apply_to_all_fields=True)
                if student_semester_code:
                    visibility_filter &= (Q(target_semester_code='') | Q(target_semester_code=student_semester_code))
                else:
                    visibility_filter &= Q(target_semester_code='')
                if student_semester_number:
                    visibility_filter &= (Q(target_semester_number__isnull=True) | Q(target_semester_number=student_semester_number))
                else:
                    visibility_filter &= Q(target_semester_number__isnull=True)
                if student_section:
                    visibility_filter &= (Q(target_section='') | Q(target_section__iexact=student_section))
                else:
                    visibility_filter &= Q(target_section='')
                qs = qs.filter(visibility_filter).distinct()
            else:
                qs = qs.none()
        else:
            qs = qs.none()

        if q:
            qs = qs.filter(
                Q(title__icontains=q)
                | Q(subject__icontains=q)
                | Q(subject_ref__name__icontains=q)
                | Q(topic_ref__name__icontains=q)
            )
        if category:
            qs = qs.filter(category__iexact=category)
        if subject:
            qs = qs.filter(Q(subject__iexact=subject) | Q(subject_ref__name__iexact=subject))
        if subject_id:
            qs = qs.filter(subject_ref_id=subject_id)
        if topic_id:
            qs = qs.filter(topic_ref_id=topic_id)
        if module_id:
            qs = qs.filter(module_ref_id=module_id)
        if semester:
            qs = qs.filter(semester=semester)
        if folder:
            qs = qs.filter(folder__iexact=folder)
        if quiz_type:
            qs = qs.filter(quiz_type=quiz_type)
        if difficulty:
            qs = qs.filter(difficulty=difficulty)
        if field_id:
            try:
                parsed_field_id = int(field_id)
                qs = qs.filter(
                    Q(assigned_fields__id=parsed_field_id) |
                    Q(target_field_of_study_id=parsed_field_id) |
                    Q(apply_to_all_fields=True)
                ).distinct()
            except (TypeError, ValueError):
                pass
        if semester_code:
            qs = qs.filter(target_semester_code=semester_code)
        if is_published in {'1', 'true', 'yes'}:
            qs = qs.filter(is_published=True)
        elif is_published in {'0', 'false', 'no'}:
            qs = qs.filter(is_published=False)
        if semester_number:
            qs = qs.filter(target_semester_number=semester_number)
        if section:
            qs = qs.filter(target_section__iexact=section)
        if visibility in {Quiz.VISIBILITY_PUBLIC, Quiz.VISIBILITY_PRIVATE}:
            qs = qs.filter(visibility=visibility)
        if min_questions:
            try:
                qs = qs.filter(question_count__gte=int(min_questions))
            except (TypeError, ValueError):
                pass
        if mine in {'1', 'true', 'yes'}:
            if request.user.is_authenticated:
                qs = qs.filter(owner=request.user)
            else:
                qs = qs.none()

        if sort in {'new', 'newest', '-created_at'}:
            qs = qs.order_by('-created_at')
        elif sort in {'old', 'oldest', 'created_at'}:
            qs = qs.order_by('created_at')
        elif sort in {'title', 'a-z'}:
            qs = qs.order_by('title')
        elif sort in {'-title', 'z-a'}:
            qs = qs.order_by('-title')
        elif sort in {'questions', 'question_count'}:
            qs = qs.order_by('-question_count', '-created_at')

        return qs

    def perform_create(self, serializer):
        quiz = serializer.save(owner=self.request.user)
        user_field_id = getattr(self.request.user, 'field_of_study_id', None)
        if not quiz.apply_to_all_fields and not quiz.assigned_fields.exists():
            if quiz.subject_ref_id:
                quiz.assigned_fields.set([quiz.subject_ref.field_of_study_id])
            elif user_field_id:
                quiz.assigned_fields.set([user_field_id])

    def get_permissions(self):
        if self.action in ['create', 'update', 'partial_update', 'destroy', 'publish', 'add_bank_questions']:
            permission_classes = [permissions.IsAuthenticated, IsTeacherOrOwner]
        else:
            permission_classes = [permissions.IsAuthenticated]
        return [p() for p in permission_classes]

    @action(detail=True, methods=['post'], url_path='publish')
    def publish(self, request, pk=None):
        quiz = self.get_object()
        self.check_object_permissions(request, quiz)
        publish_flag = request.data.get('is_published', True)
        if isinstance(publish_flag, str):
            publish_flag = publish_flag.strip().lower() in {'1', 'true', 'yes', 'on'}
        quiz.is_published = bool(publish_flag)
        quiz.save(update_fields=['is_published', 'updated_at'])
        return Response({'id': quiz.id, 'is_published': quiz.is_published})

    @action(detail=True, methods=['post'], url_path='add-bank-questions')
    def add_bank_questions(self, request, pk=None):
        quiz = self.get_object()
        self.check_object_permissions(request, quiz)
        question_ids = request.data.get('question_ids') or []
        if not isinstance(question_ids, list) or not question_ids:
            return Response({'detail': 'question_ids must be a non-empty list.'}, status=400)

        entries = list(
            QuestionBank.objects.filter(id__in=question_ids, is_active=True).select_related('subject_ref', 'topic_ref', 'module_ref')
        )
        if not entries:
            return Response({'detail': 'No valid question bank entries found.'}, status=400)

        start_order = quiz.questions.count()
        created_ids = []
        for offset, entry in enumerate(entries):
            question = Question.objects.create(
                quiz=quiz,
                subject_ref=entry.subject_ref,
                topic_ref=entry.topic_ref,
                module_ref=entry.module_ref,
                text=entry.question_text,
                explanation=entry.explanation,
                question_type=entry.question_type,
                difficulty=entry.difficulty,
                marks=entry.marks,
                timer_seconds=30,
                order=start_order + offset,
            )
            options = [entry.option_a, entry.option_b, entry.option_c, entry.option_d]
            normalized_correct = (entry.correct_answer or '').strip().upper()
            if question.question_type == Question.TYPE_TRUE_FALSE:
                options = ['True', 'False']
            for index, option in enumerate(options):
                if not option:
                    continue
                is_correct = normalized_correct == chr(65 + index) or option.upper() == normalized_correct
                question.choices.create(text=option, is_correct=is_correct, order=index)
            QuizQuestion.objects.create(
                quiz=quiz,
                question=question,
                question_bank_reference=entry,
                order=question.order,
            )
            created_ids.append(question.id)
        return Response({'created_question_ids': created_ids}, status=status.HTTP_201_CREATED)


class QuestionBankBulkImportSerializer(serializers.Serializer):
    subject_ref = serializers.PrimaryKeyRelatedField(queryset=Subject.objects.filter(is_active=True))
    topic_ref = serializers.PrimaryKeyRelatedField(queryset=Topic.objects.filter(is_active=True), required=False, allow_null=True)
    module_ref = serializers.PrimaryKeyRelatedField(queryset=Module.objects.filter(is_active=True), required=False, allow_null=True)
    field_of_study = serializers.IntegerField(required=False)
    semester = serializers.IntegerField(required=False, min_value=1)
    format = serializers.ChoiceField(choices=['csv', 'json', 'excel'], required=False)
    preview = serializers.BooleanField(required=False, default=True)
    content = serializers.CharField(required=False, allow_blank=True)


class QuestionBankViewSet(viewsets.ModelViewSet):
    serializer_class = QuestionBankSerializer

    def get_queryset(self):
        qs = QuestionBank.objects.select_related(
            'subject_ref', 'topic_ref', 'module_ref', 'field_of_study', 'created_by'
        ).order_by('-updated_at')
        user = self.request.user
        if not user.is_admin():
            qs = qs.filter(created_by=user)

        subject_id = (self.request.query_params.get('subject_id') or '').strip()
        topic_id = (self.request.query_params.get('topic_id') or '').strip()
        module_id = (self.request.query_params.get('module_id') or '').strip()
        field_id = (self.request.query_params.get('field_id') or '').strip()
        semester = (self.request.query_params.get('semester') or '').strip()
        difficulty = (self.request.query_params.get('difficulty') or '').strip()
        question_type = (self.request.query_params.get('question_type') or '').strip()
        query = (self.request.query_params.get('q') or '').strip()

        if subject_id:
            qs = qs.filter(subject_ref_id=subject_id)
        if topic_id:
            qs = qs.filter(topic_ref_id=topic_id)
        if module_id:
            qs = qs.filter(module_ref_id=module_id)
        if field_id:
            qs = qs.filter(field_of_study_id=field_id)
        if semester:
            qs = qs.filter(semester=semester)
        if difficulty:
            qs = qs.filter(difficulty=difficulty)
        if question_type:
            qs = qs.filter(question_type=question_type)
        if query:
            qs = qs.filter(
                Q(question_text__icontains=query)
                | Q(subject_ref__name__icontains=query)
                | Q(topic__icontains=query)
                | Q(unit_name__icontains=query)
            )
        return qs

    def get_permissions(self):
        return [permissions.IsAuthenticated(), IsTeacherOrOwner()]

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    @staticmethod
    def _infer_format(serializer, upload):
        declared = serializer.validated_data.get('format')
        if declared:
            return declared
        if upload is None:
            return 'json'
        suffix = Path(upload.name).suffix.lower()
        if suffix == '.csv':
            return 'csv'
        if suffix in {'.xlsx', '.xlsm', '.xltx', '.xltm'}:
            return 'excel'
        return 'json'

    @staticmethod
    def _row_to_payload(row, base_payload):
        payload = {
            'subject_ref': base_payload['subject_ref'].id,
            'topic_ref': base_payload['topic_ref'].id if base_payload.get('topic_ref') else None,
            'module_ref': base_payload['module_ref'].id if base_payload.get('module_ref') else None,
            'field_of_study': base_payload.get('field_of_study') or base_payload['subject_ref'].field_of_study_id,
            'semester': base_payload.get('semester') or base_payload['subject_ref'].semester,
            'topic': str(row.get('topic') or getattr(base_payload.get('topic_ref'), 'name', '') or '').strip(),
            'unit_name': str(row.get('unit') or row.get('unit_name') or getattr(base_payload.get('topic_ref'), 'unit_name', '') or '').strip(),
            'question_text': str(row.get('question_text') or row.get('question') or '').strip(),
            'question_type': str(row.get('question_type') or row.get('type') or Question.TYPE_MCQ).strip() or Question.TYPE_MCQ,
            'difficulty': str(row.get('difficulty') or Question.DIFFICULTY_MEDIUM).strip().lower(),
            'option_a': str(row.get('option_a') or '').strip(),
            'option_b': str(row.get('option_b') or '').strip(),
            'option_c': str(row.get('option_c') or '').strip(),
            'option_d': str(row.get('option_d') or '').strip(),
            'correct_answer': str(row.get('correct_answer') or row.get('answer') or '').strip(),
            'explanation': str(row.get('explanation') or '').strip(),
            'marks': int(row.get('marks') or 1),
        }
        if payload['question_type'] in {'true_false', 'true/false', 'tf'}:
            payload['question_type'] = Question.TYPE_TRUE_FALSE
            payload['option_a'] = 'True'
            payload['option_b'] = 'False'
        return payload

    def _parse_import_rows(self, import_format, upload, content):
        if import_format == 'csv':
            text = content
            if upload is not None:
                text = upload.read().decode('utf-8-sig')
            return list(csv.DictReader(io.StringIO(text)))
        if import_format == 'json':
            text = content
            if upload is not None:
                text = upload.read().decode('utf-8-sig')
            data = json.loads(text or '[]')
            if not isinstance(data, list):
                raise serializers.ValidationError('JSON import must be a list of question objects.')
            return data
        if import_format == 'excel':
            try:
                from openpyxl import load_workbook
            except ImportError as exc:  # pragma: no cover - environment dependent
                raise serializers.ValidationError('openpyxl is required for Excel import.') from exc
            if upload is None:
                raise serializers.ValidationError('Excel import requires a file upload.')
            workbook = load_workbook(upload, read_only=True)
            worksheet = workbook.active
            rows = list(worksheet.iter_rows(values_only=True))
            if not rows:
                return []
            headers = [str(value or '').strip() for value in rows[0]]
            return [
                {headers[index]: row[index] for index in range(len(headers))}
                for row in rows[1:]
            ]
        raise serializers.ValidationError('Unsupported import format.')

    @action(detail=False, methods=['post'], url_path='bulk-import')
    def bulk_import(self, request):
        serializer = QuestionBankBulkImportSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        upload = request.FILES.get('file')
        import_format = self._infer_format(serializer, upload)
        rows = self._parse_import_rows(import_format, upload, serializer.validated_data.get('content', ''))
        if not rows:
            return Response({'detail': 'No rows found in import payload.'}, status=400)

        preview = serializer.validated_data.get('preview', True)
        created = []
        failures = []
        for row_number, row in enumerate(rows, start=2):
            payload = self._row_to_payload(row, serializer.validated_data)
            item_serializer = self.get_serializer(data=payload, context={'request': request})
            if item_serializer.is_valid():
                if not preview:
                    created.append(item_serializer.save(created_by=request.user))
            else:
                failures.append({'row': row_number, 'errors': item_serializer.errors})

        return Response(
            {
                'preview': preview,
                'format': import_format,
                'rows_received': len(rows),
                'valid_rows': len(rows) - len(failures),
                'created_count': 0 if preview else len(created),
                'failed_rows': failures,
                'created_ids': [] if preview else [item.id for item in created],
            },
            status=status.HTTP_200_OK,
        )

    @action(detail=False, methods=['post'], url_path='import')
    def import_alias(self, request):
        return self.bulk_import(request)

    @action(detail=True, methods=['post'], url_path='duplicate')
    def duplicate(self, request, pk=None):
        entry = self.get_object()
        self.check_object_permissions(request, entry)
        duplicated = QuestionBank.objects.create(
            subject_ref=entry.subject_ref,
            topic_ref=entry.topic_ref,
            module_ref=entry.module_ref,
            field_of_study=entry.field_of_study,
            semester=entry.semester,
            topic=entry.topic,
            unit_name=entry.unit_name,
            question_text=entry.question_text,
            question_type=entry.question_type,
            difficulty=entry.difficulty,
            option_a=entry.option_a,
            option_b=entry.option_b,
            option_c=entry.option_c,
            option_d=entry.option_d,
            correct_answer=entry.correct_answer,
            explanation=entry.explanation,
            marks=entry.marks,
            created_by=request.user,
        )
        return Response(self.get_serializer(duplicated).data, status=status.HTTP_201_CREATED)
